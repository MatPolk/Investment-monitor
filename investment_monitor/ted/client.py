"""
TED EU Open Data API client.

Fetches construction tenders from the EU's TED (Tenders Electronic Daily)
for Poland (CPV 45*) above the configured value threshold.

Pagination strategy (mirrors Kompas):
  - Default run: fetch from the date saved in ted_last_date.txt
  - First run (no state file): fetch the last 365 days
  - --days-back N override: explicit date range (useful for backfill)

After a successful fetch the newest publication date is saved so the
next run picks up exactly where this one left off (minus a 1-day safety
margin to handle API propagation delays).
"""
import logging
import difflib
from datetime import datetime, timedelta
from time import sleep

import requests

from investment_monitor.config import TED_API_URL, TED_MIN_VALUE_PLN, TED_MIN_VALUE_NETTO
from investment_monitor.state import load_ted_last_date, save_ted_last_date
from investment_monitor.ted.normalizer import normalize_ted_name, TED_NOTICE_STATUS, nuts_to_woj, is_construction
from investment_monitor.utils.text import clean_company_name
from investment_monitor.classify.classifier import classify_investment


# ── Internal: parse one TED API record ───────────────────────────────────────

def _parse_item(item: dict) -> dict:
    """Map one TED API v3 record to a dict matching the database column layout."""

    # Title — prefer title-proc (clean Polish name without "Polska — " prefix)
    title_proc   = item.get("title-proc") or {}
    title_lot    = item.get("title-lot") or {}
    notice_title = item.get("notice-title") or {}

    if isinstance(title_proc, dict) and (title_proc.get("pol") or title_proc.get("eng")):
        name = title_proc.get("pol") or title_proc.get("eng") or ""
    elif isinstance(title_lot, dict):
        lots = title_lot.get("pol") or title_lot.get("eng") or []
        name = lots[0] if isinstance(lots, list) and lots else str(title_lot)
    elif isinstance(notice_title, dict):
        name = notice_title.get("pol") or notice_title.get("eng") or ""
    else:
        name = str(notice_title)

    # Description (fallback for rail line number extraction)
    desc_obj    = item.get("description-lot") or {}
    if isinstance(desc_obj, dict):
        desc_list   = desc_obj.get("pol") or desc_obj.get("eng") or []
        description = desc_list[0] if isinstance(desc_list, list) and desc_list else ""
    elif isinstance(desc_obj, list):
        description = desc_obj[0] if desc_obj else ""
    else:
        description = str(desc_obj) if desc_obj else ""

    # Buyer / investor
    auth_obj = item.get("organisation-name-buyer") or {}
    if isinstance(auth_obj, dict):
        inv_list = auth_obj.get("pol") or auth_obj.get("eng") or []
        investor = inv_list[0] if isinstance(inv_list, list) and inv_list else ""
    else:
        investor = str(auth_obj)

    # Winner / contractor
    winner_obj = item.get("winner-name") or {}
    if isinstance(winner_obj, dict):
        w_list = winner_obj.get("pol") or winner_obj.get("eng") or []
        winner = w_list[0] if isinstance(w_list, list) and w_list else ""
    else:
        winner = str(winner_obj) if winner_obj else ""

    def _smart_title(s: str) -> str:
        """Title-case an ALL-CAPS string while preserving short acronyms (PKP, PGE)."""
        return " ".join(w if (len(w) <= 4 and w.isupper()) else w.capitalize() for w in s.split())

    investor = clean_company_name(investor)
    winner   = clean_company_name(winner)
    if investor and investor.isupper() and len(investor) > 3:
        investor = _smart_title(investor)
    if winner and winner.isupper() and len(winner) > 3:
        winner = _smart_title(winner)

    # Location
    cities    = item.get("place-of-performance-city-lot") or []
    city      = cities[0] if cities else ""
    nuts_list = item.get("place-of-performance-subdiv-lot") or []
    nuts_code = nuts_list[0] if nuts_list else ""
    woj       = nuts_to_woj(nuts_code)

    # Value: contract > total > estimated; convert PLN (gross→net) or EUR
    value_mln = None
    for val_field, cur_field in [
        ("tender-value",      "tender-value-cur"),
        ("total-value",       "total-value-cur"),
        ("estimated-value-lot","estimated-value-cur-lot"),
    ]:
        value_raw = item.get(val_field)
        if value_raw is None:
            continue
        if isinstance(value_raw, list):
            value_raw = value_raw[0] if value_raw else None
        if value_raw is None:
            continue
        currencies = item.get(cur_field) or []
        currency   = (currencies[0] if isinstance(currencies, list) and currencies
                      else str(currencies) or "PLN")
        try:
            v = float(value_raw)
            if any(t in currency.upper() for t in ("PLN", "ZLOTY", "ZL")):
                value_mln = round(v / 1_000_000 / 1.23, 1)
            else:
                value_mln = round(v * 4.25 / 1_000_000 / 1.23, 1)
        except (TypeError, ValueError):
            continue
        break

    # Notice type → status
    notice_type = item.get("notice-type") or ""
    status      = TED_NOTICE_STATUS.get(notice_type, "Przetarg")

    # Publication date
    pub_date_raw = item.get("publication-date") or ""
    if isinstance(pub_date_raw, list):
        pub_date_raw = pub_date_raw[0] if pub_date_raw else ""
    pub_date = str(pub_date_raw)[:10]
    if len(pub_date) == 8 and pub_date.isdigit():
        pub_date = f"{pub_date[:4]}-{pub_date[4:6]}-{pub_date[6:8]}"

    pub_num = item.get("publication-number") or ""
    link    = f"https://ted.europa.eu/pl/notice/-/detail/{pub_num}" if pub_num else ""

    norm_name, norm_city = normalize_ted_name(str(name), city, description)
    if norm_city and not city:
        city = norm_city

    cls = classify_investment(norm_name, investor)

    return {
        "Inwestycja":           norm_name,
        "Miejscowość":          norm_city or city,
        "Województwo":          woj,
        "Inwestor":             investor,
        "Generalny wykonawca":  winner,
        "Status inwestycji":    status,
        "Wartość (mln zł)":     value_mln,
        "Sektor":               cls["sektor"] or "",
        "Sektor.1":             cls["sektor_pkob"] or "",
        "Znaczące segmenty":    cls["segment"] or "",
        "Remonty, modernizacje, rewitalizacje, przebudowy": cls["remonty"] or "",
        "Inwestycje wojskowe":  cls["wojsko"] or "",
        "Linki":                link,
        "_source":              "TED",
        "_notice_type":         notice_type,
        "_pub_num":             pub_num,
        "_pub_date":            pub_date,
    }


# ── Public API ────────────────────────────────────────────────────────────────

def fetch_ted(df_baza, days_back: int | None = None) -> tuple[list, list]:
    """
    Fetch TED construction tenders for Poland above the configured value threshold.

    Date range logic:
      - days_back=N  → explicit override (backfill / re-scan)
      - days_back=None, state file exists  → from last saved date − 1 day
      - days_back=None, no state file      → last 365 days (first run)

    Returns (new_items, matched_items):
      new_items     — notices not found in df_baza
      matched_items — notices fuzzy-matched to existing rows in df_baza
    """
    from investment_monitor.matching.fuzzy import fuzzy_match_ted  # local import avoids circular dep
    from openpyxl import load_workbook
    import os
    from investment_monitor.config import OUTPUT_FILE

    if days_back is not None:
        date_from = (datetime.now() - timedelta(days=days_back)).strftime("%Y%m%d")
        logging.info(f"TED: manual override --days-back {days_back}, from {date_from}")
    else:
        last_date = load_ted_last_date()
        if last_date:
            safe = (datetime.strptime(last_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y%m%d")
            date_from = safe
            logging.info(f"TED: auto mode, from last date {last_date} (with 1-day margin: {safe})")
        else:
            date_from = (datetime.now() - timedelta(days=365)).strftime("%Y%m%d")
            logging.info(f"TED: first run — no state file — scanning from {date_from}")

    all_items: list[dict] = []
    page = 1

    while True:
        try:
            logging.info(f"TED: page {page}, CPV 45*, estimated >= {TED_MIN_VALUE_NETTO:.0f}M PLN net, from {date_from}...")
            body = {
                "query": (
                    f"organisation-country-buyer=POL "
                    f"AND classification-cpv=45* "
                    f"AND publication-date>={date_from} "
                    f"AND estimated-value-lot>={TED_MIN_VALUE_PLN}"
                ),
                "fields": [
                    "publication-number", "notice-type", "publication-date",
                    "title-proc", "title-lot", "notice-title", "description-lot",
                    "estimated-value-lot", "estimated-value-cur-lot",
                    "tender-value", "tender-value-cur",
                    "total-value", "total-value-cur",
                    "organisation-name-buyer", "winner-name",
                    "place-of-performance-city-lot",
                    "place-of-performance-subdiv-lot",
                    "classification-cpv",
                ],
                "page": page, "limit": 100, "onlyLatestVersions": True,
            }
            resp = requests.post(TED_API_URL, json=body, timeout=30)
            resp.raise_for_status()
            data    = resp.json()
            notices = data.get("notices") or []
            total   = data.get("totalNoticeCount", 0)
            logging.info(f"TED: page {page} → {len(notices)} results (total: {total})")

            for notice in notices:
                raw_title = ""
                for tf in ("title-proc", "title-lot", "notice-title"):
                    t = notice.get(tf) or {}
                    raw_title = (t.get("pol") or t.get("eng") or "") if isinstance(t, dict) else str(t)
                    if raw_title:
                        break
                if isinstance(raw_title, list):
                    raw_title = raw_title[0] if raw_title else ""
                if not is_construction(str(raw_title)):
                    continue
                all_items.append(_parse_item(notice))

            if len(notices) < 100:
                break
            page += 1
            sleep(0.5)
        except Exception as e:
            logging.warning(f"TED: error on page {page} ({e})")
            break

    if not all_items:
        logging.info("TED: no results")
        return [], []

    # Persist the newest publication date seen
    newest_date = max((item.get("_pub_date", "") for item in all_items), default="")
    if newest_date:
        save_ted_last_date(newest_date)
        logging.info(f"TED: saved last date → {newest_date}")

    logging.info(f"TED: {len(all_items)} notices (before value filter)")

    # Post-filter: net value >= threshold (API uses gross estimated; we need net)
    all_items = [
        item for item in all_items
        if item.get("Wartość (mln zł)") is None or item["Wartość (mln zł)"] >= TED_MIN_VALUE_NETTO
    ]
    logging.info(f"TED: {len(all_items)} notices after net-value filter (>= {TED_MIN_VALUE_NETTO}M PLN)")

    # Deduplicate by publication number
    seen_pubs, unique = set(), []
    for item in all_items:
        pub = item.get("_pub_num", "")
        if pub and pub in seen_pubs:
            continue
        if pub:
            seen_pubs.add(pub)
        unique.append(item)

    # Deduplicate by normalised name (prefer 'result' over 'can', prefer newer date)
    seen_names: dict[str, dict] = {}
    deduped = []
    for item in unique:
        key = item.get("Inwestycja", "").lower().strip()
        if not key:
            deduped.append(item)
            continue
        if key in seen_names:
            ex = seen_names[key]
            if ("result" in item.get("_notice_type", "") and "result" not in ex.get("_notice_type", "")) \
               or item.get("_pub_date", "") > ex.get("_pub_date", ""):
                seen_names[key] = item
                deduped = [x for x in deduped if x.get("Inwestycja", "").lower().strip() != key]
                deduped.append(item)
        else:
            seen_names[key] = item
            deduped.append(item)
    unique = deduped

    # Collect known TED links (skip already-processed notices)
    link_cols = [c for c in df_baza.columns if "Linki" in c or c.startswith("Unnamed")]
    known_links: set[str] = set()
    for _, row in df_baza.iterrows():
        for col in link_cols:
            v = str(row.get(col, "")).strip()
            if v and "ted.europa.eu" in v:
                known_links.add(v)
    if os.path.exists(OUTPUT_FILE):
        try:
            wb = load_workbook(OUTPUT_FILE, read_only=True)
            for sn in wb.sheetnames:
                for row in wb[sn].iter_rows(min_row=2, values_only=True):
                    for val in row:
                        v = str(val).strip() if val else ""
                        if v and "ted.europa.eu" in v:
                            known_links.add(v)
            wb.close()
        except Exception:
            pass

    new_items, matched_items = [], []
    for item in unique:
        link = item.get("Linki", "")
        if link and link in known_links:
            continue
        match_idx, score, match_type = fuzzy_match_ted(item, df_baza)
        if match_idx is not None:
            item["_match_idx"]   = match_idx
            item["_match_score"] = score
            item["_match_type"]  = match_type
            matched_items.append(item)
        else:
            new_items.append(item)

    logging.info(f"TED: {len(new_items)} new, {len(matched_items)} matched to database")
    return new_items, matched_items

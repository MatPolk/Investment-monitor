"""
Excel output generation.

Produces do_zatwierdzenia.xlsx with two sheets:

  "Nowe"       — new investments from Kompas and TED combined
  "Dopasowane" — existing database rows with proposed field updates
                 highlighted in green

Previously generated rows are preserved and appended below new ones
on each run, so the analyst can review changes from multiple runs
in a single file.
"""
import os
import logging
from datetime import date

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from investment_monitor.config import DATABASE_FILE, OUTPUT_FILE
from investment_monitor.kompas.scraper import extract_id


# ── Style constants ───────────────────────────────────────────────────────────

FILL_GREEN  = PatternFill("solid", fgColor="92D050")
FONT_HEADER = Font(bold=True)

_STATUS_FONT = {
    "Budowa":                "FF000000",
    "Planowanie":            "FF000000",
    "Przetarg":              "FF000000",
    "Wstępna koncepcja":     "FF009999",
    "Inwestycja wstrzymana": "FF7030A0",
    "Inwestycja zakończona": "FFFF0000",
    "Inwestycja zarzucona":  "FF806000",
}
_URL_COLS = set(range(26, 33))

_F_YELLOW = PatternFill("solid", fgColor="FFFFFF00")
_F_CREAM  = PatternFill("solid", fgColor="FFFFF2CC")
_F_BLUE   = PatternFill("solid", fgColor="FFB7DEE8")
_F_ORANGE = PatternFill("solid", fgColor="FFFFC000")

CELL_FILLS = {
    1: _F_CREAM, 3: _F_YELLOW, 7: _F_YELLOW, 8: _F_YELLOW,
    11: _F_YELLOW, 12: _F_YELLOW, 16: _F_YELLOW, 18: _F_YELLOW,
    21: _F_BLUE, 22: _F_BLUE, 23: _F_BLUE, 24: _F_BLUE, 25: _F_BLUE,
    26: _F_YELLOW, 27: _F_YELLOW, 28: _F_YELLOW, 29: _F_YELLOW,
    30: _F_YELLOW, 31: _F_YELLOW, 32: _F_YELLOW,
    33: _F_ORANGE,
}


# ── Low-level write helpers ───────────────────────────────────────────────────

def write_header(ws, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        if col_idx in CELL_FILLS:
            cell.fill = CELL_FILLS[col_idx]
        cell.font      = FONT_HEADER
        cell.alignment = Alignment(horizontal="left")


def write_data_row(ws, row_num, values, font_rgb="FF000000", green_cols=None, orange_col1=False):
    font_hex = font_rgb[2:] if font_rgb.startswith("FF") else font_rgb
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal="left", wrap_text=False)
        if col_idx == 1 and orange_col1:
            cell.fill = PatternFill("solid", fgColor="FFB347")
        elif green_cols and col_idx in green_cols:
            cell.fill = FILL_GREEN
        elif col_idx in CELL_FILLS:
            cell.fill = CELL_FILLS[col_idx]
        if val and str(val).startswith("http"):
            cell.hyperlink = str(val)
            cell.font = Font(color="0563C1", underline="single")
        else:
            cell.font = Font(color=font_hex)


def apply_sheet_settings(ws, col_widths, n_cols):
    ws.sheet_view.zoomScale = 70
    ws.freeze_panes         = "B2"
    if n_cols > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"
    for i in range(1, n_cols + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = col_widths.get(letter, 18)
    ws.column_dimensions["A"].width = 16


# ── Database formatting ───────────────────────────────────────────────────────

def load_baza_formatting():
    """Read row font colours and column widths from the reference database."""
    wb = load_workbook(DATABASE_FILE)
    ws = wb["Baza"]
    row_fonts = {}
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        rgb  = "FF000000"
        if cell.font and cell.font.color:
            try:
                raw = cell.font.color.rgb
                if raw and isinstance(raw, str):
                    rgb = raw
            except Exception:
                pass
        row_fonts[cell.row - 2] = rgb
    col_widths = {l: cd.width for l, cd in ws.column_dimensions.items() if cd.width}
    wb.close()
    logging.info(f"Database formatting loaded ({len(row_fonts)} rows)")
    return row_fonts, col_widths


# ── Previous-run row preservation ────────────────────────────────────────────

def load_existing_rows() -> dict:
    """Load rows from a previous run of do_zatwierdzenia.xlsx."""
    if not os.path.exists(OUTPUT_FILE):
        return {}
    try:
        wb = load_workbook(OUTPUT_FILE)
    except Exception as e:
        logging.warning(f"Could not load existing output file: {e}")
        return {}

    result = {}
    for sheet_name in wb.sheetnames:
        ws, rows = wb[sheet_name], []
        for row_idx in range(2, ws.max_row + 1):
            values, kompas_id, font_rgb, orange, green_cols = [], None, "FF000000", False, set()
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, c)
                val  = cell.value
                values.append(str(val) if val is not None else "")
                if c in _URL_COLS and val:
                    kid = extract_id(str(val))
                    if kid:
                        kompas_id = kid
                if c == 1:
                    try:
                        raw = cell.font.color.rgb if cell.font and cell.font.color else None
                        if isinstance(raw, str) and raw:
                            font_rgb = raw
                    except Exception:
                        pass
                    try:
                        fc = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
                        if isinstance(fc, str) and "FFB347" in fc:
                            orange = True
                    except Exception:
                        pass
                try:
                    fc = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
                    if isinstance(fc, str) and "92D050" in fc:
                        green_cols.add(c)
                except Exception:
                    pass
            if any(v.strip() for v in values):
                rows.append({"values": values, "kompas_id": kompas_id,
                             "font_rgb": font_rgb, "orange": orange, "green_cols": green_cols})
        result[sheet_name] = rows
    return result


# ── Main output generator ─────────────────────────────────────────────────────

def generate_output(nowe, df_baza, row_fonts, col_widths,
                    ted_nowe=None, ted_dopasowane=None, kompas_dopasowane=None):
    """Write do_zatwierdzenia.xlsx with 'Nowe' and 'Dopasowane' sheets."""
    existing_rows  = load_existing_rows()
    ted_nowe       = ted_nowe or []
    ted_dopasowane = ted_dopasowane or []
    kompas_dopasowane = kompas_dopasowane or []

    current_ids = {extract_id(item.get("url", "")) for item in nowe} - {None}

    def old_rows_for(sheet):
        return [r for r in existing_rows.get(sheet, []) if r.get("kompas_id") not in current_ids]

    baza_cols = list(df_baza.columns)
    rest_cols = baza_cols[1:]  # skip "Inwestycja zbiorcza" → replaced by "Data dodania"
    headers   = ["Data dodania"] + rest_cols
    n_cols    = len(headers)
    today_str = date.today().strftime("%d.%m.%Y")

    val_col_idx = headers.index("Wartość (mln zł)") if "Wartość (mln zł)" in headers else None

    def _val(data_dict, col_name):
        val = data_dict.get(col_name, "")
        if col_name == "Wartość (mln zł)":
            if val is None or val == "":
                return None
            try:
                return float(str(val).replace(",", "."))
            except (ValueError, TypeError):
                return None
        return str(val) if val is not None else ""

    def append_old(ws, old_rows, row_num, nc):
        for old in old_rows:
            vals = list(old["values"])
            if len(vals) < nc:
                vals += [""] * (nc - len(vals))
            if val_col_idx is not None and val_col_idx < len(vals):
                v = vals[val_col_idx]
                if v and v not in ("", "None"):
                    try:
                        vals[val_col_idx] = float(str(v).replace(",", "."))
                    except (ValueError, TypeError):
                        pass
                elif v in ("", "None"):
                    vals[val_col_idx] = None
            write_data_row(ws, row_num, vals[:nc], old["font_rgb"], old.get("green_cols"), old["orange"])
            row_num += 1
        return row_num

    wb = Workbook()

    # ── Sheet 1: Nowe ─────────────────────────────────────────────────────────
    ws_nowe = wb.active
    ws_nowe.title = "Nowe"
    write_header(ws_nowe, headers)
    row_num = 2
    for item in nowe:
        kd     = item["kompas_data"]
        status = kd.get("Status inwestycji") or ""
        write_data_row(ws_nowe, row_num, [today_str] + [_val(kd, c) for c in rest_cols],
                       font_rgb=_STATUS_FONT.get(status, "FF000000"))
        row_num += 1
    for item in ted_nowe:
        status = item.get("Status inwestycji") or ""
        write_data_row(ws_nowe, row_num, [item.get("_pub_date", today_str)] + [_val(item, c) for c in rest_cols],
                       font_rgb=_STATUS_FONT.get(status, "FF000000"))
        row_num += 1
    append_old(ws_nowe, old_rows_for("Nowe"), row_num, n_cols)
    apply_sheet_settings(ws_nowe, col_widths, n_cols)

    # ── Sheet 2: Dopasowane ───────────────────────────────────────────────────
    dopas_headers = ["Link źródła"] + rest_cols + ["Match score", "Match type"]
    n_cols_d      = len(dopas_headers)
    ws_dopas      = wb.create_sheet("Dopasowane")
    write_header(ws_dopas, dopas_headers)
    row_num = 2

    for item in ted_dopasowane:
        baza_idx = item.get("_match_idx")
        baza_row = df_baza.loc[baza_idx] if baza_idx is not None else None
        values, green_cols = [item.get("Linki", "")], set()
        for ci, c in enumerate(rest_cols, 2):
            baza_val  = baza_row.get(c, "") if baza_row is not None else ""
            ted_val   = item.get(c, "")
            disp_val  = _val({"v": baza_val}, "v") if c != "Wartość (mln zł)" else None
            if c == "Wartość (mln zł)":
                try:
                    disp_val = float(str(baza_val).replace(",", ".")) if baza_val not in (None, "", "nan") else None
                except (ValueError, TypeError):
                    disp_val = None
            if ted_val and str(ted_val).strip():
                baza_str = str(baza_val).strip() if baza_val not in (None, "", "nan") else ""
                ted_str  = str(ted_val).strip()
                differs  = baza_str != ted_str
                if c == "Wartość (mln zł)":
                    try:
                        differs = abs(float(str(baza_val or 0).replace(",", ".")) - float(str(ted_val).replace(",", "."))) > 0.5
                    except (ValueError, TypeError):
                        pass
                if not baza_str or (differs and c in ("Wartość (mln zł)", "Generalny wykonawca", "Miejscowość", "Województwo")):
                    disp_val = _val(item, c)
                    green_cols.add(ci)
            values.append(disp_val)
        score = item.get("_match_score")
        values += [f"{score:.0%}" if score is not None else "", item.get("_match_type", "")]
        font_rgb = row_fonts.get(baza_idx, "FF000000") if baza_idx is not None else "FF000000"
        write_data_row(ws_dopas, row_num, values, font_rgb=font_rgb, green_cols=green_cols)
        row_num += 1

    for item in kompas_dopasowane:
        baza_idx  = item.get("baza_row_idx")
        baza_row  = item.get("baza_row")
        kompas_data = item.get("kompas_data", {})
        values, green_cols = [kompas_data.get("Linki", "")], set()
        for ci, c in enumerate(rest_cols, 2):
            baza_val = baza_row.get(c, "") if baza_row is not None else ""
            disp_val = _val({"v": baza_val}, "v") if c != "Wartość (mln zł)" else None
            if c == "Wartość (mln zł)":
                try:
                    disp_val = float(str(baza_val).replace(",", ".")) if baza_val not in (None, "", "nan") else None
                except (ValueError, TypeError):
                    disp_val = None
            kompas_val = kompas_data.get(c, "")
            if kompas_val and str(kompas_val).strip() and not str(baza_val).strip():
                disp_val = _val(kompas_data, c)
                green_cols.add(ci)
            values.append(disp_val)
        score = item.get("fuzzy_score")
        values += [f"{score:.0%}" if score is not None else "", item.get("match_type", "KOMPAS")]
        font_rgb = row_fonts.get(baza_idx, "FF000000") if baza_idx is not None else "FF000000"
        write_data_row(ws_dopas, row_num, values, font_rgb=font_rgb, green_cols=green_cols)
        row_num += 1

    append_old(ws_dopas, old_rows_for("Dopasowane"), row_num, n_cols_d)
    apply_sheet_settings(ws_dopas, col_widths, n_cols_d)

    wb.save(OUTPUT_FILE)
    logging.info(
        f"Saved: {OUTPUT_FILE} | New: {len(nowe)} Kompas + {len(ted_nowe)} TED "
        f"| Matched: {len(ted_dopasowane)} TED + {len(kompas_dopasowane)} Kompas"
    )
